import tkinter as tk
from tkinter import ttk
import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os


def make_request_with_retries(url, retries=5, delay=100):
    for attempt in range(retries):
        try:
            response = requests.get(url)
            response.raise_for_status()
            return response
        except requests.exceptions.RequestException as e:
            print(f"Ошибка подключения: {e} - Попытка {attempt + 1} из {retries}")
            if attempt == retries - 1:
                return None
            time.sleep(delay)
    return None

def collect_links(base_url, day, month):
    links = set()
    a = 0
    url = base_url + str(month) + "-" + str(day) + "/"
    response = make_request_with_retries(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    holidays_items = soup.find_all('div', class_='caption')
    for item in holidays_items:
        link = item.find_all('a')[0]['href']
        full_link = f"{link}"
        compare = "https"
        if full_link[0:5] == compare:
            links.add(full_link)
        a += 1
    return links
def collect_hd_data(links,day,month):
    hd_data = set()
    for link in links:
        print(link)
        response = make_request_with_retries(link)
        if response is None:
            continue
        soup = BeautifulSoup(response.content, 'html.parser')
        title_tag = soup.find('h1')
        title = title_tag.get_text(strip=True)
        if (int(month) > 9):
            date = str(day) + "." + str(month)
        else:
            date = str(day) + ".0" + str(month)
        type_tag = soup.find_all('span', itemprop='name')[1]
        typo = type_tag.get_text(strip=True)
        type_str = typo if typo else "Нет типа"
        hd_data.add((title, link, date, type_str))
    return hd_data

def save_data_to_sheet(ws, data, old_data):
    ws.delete_rows(2, ws.max_row)
    for title, link, date, type_str in data:
        status = "Old"
        if (title, link, date, type_str) not in old_data:
            status = "New"
        ws.append([title, link, date, type_str, status])

sections = {
    "Общее": "https://www.calend.ru/day/",
    "Праздники": "https://www.calend.ru/holidays/",
    "Именины": "https://www.calend.ru/names/",
    "Народный календарь": "https://www.calend.ru/narod/",
    "Хроника": "https://www.calend.ru/events/",
    "Компании": "https://www.calend.ru/orgs/",
    "Персоны": "https://www.calend.ru/persons/"
}

file_path = "HD.xlsx"
all_hd_file_path = "ALL_HD.xlsx"
class HolidaysParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Holidays Parser")

        self.root.rowconfigure(4, weight=1)

        self.section_var = tk.StringVar()
        self.day_var = tk.StringVar()
        self.month_var = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        # Section selection
        section_label = tk.Label(self.root, text="Выберите раздел:")
        section_label.grid(row=0, column=2, columnspan=2, padx=5, pady=5, sticky="e")

        self.section_combobox = ttk.Combobox(self.root, textvariable=self.section_var, values=list(sections.keys()))
        self.section_combobox.current(0)
        self.section_combobox.grid(row=0, column=4, columnspan=1, padx=10, pady=5, sticky="nsew")

        # Page number entry
        day_label = tk.Label(self.root, text="Введите день формата д:")
        day_label.grid(row=1, column=2, columnspan=2, padx=10, pady=5, sticky="e")

        month_label = tk.Label(self.root, text="Введите месяц формата м:")
        month_label.grid(row=2, column=2, columnspan=2, padx=10, pady=5, sticky="e")

        self.day_entry = tk.Entry(self.root, textvariable=self.day_var)
        self.day_entry.grid(row=1, column=4, columnspan=1, padx=10, pady=5, sticky="nsew")

        self.month_entry = tk.Entry(self.root, textvariable=self.month_var)
        self.month_entry.grid(row=2, column=4, columnspan=1, padx=10, pady=5, sticky="nsew")

        # Parse button
        parse_button = tk.Button(self.root, text="Считать", command=self.parse_hd)
        parse_button.grid(row=0, column=0, columnspan=1, pady=10, sticky="nsew")

        # Show button
        show_button = tk.Button(self.root, text="Обновить", command=self.show_hd)
        show_button.grid(row=0, column=1, columnspan=1, pady=10, sticky="nsew")



        self.open_button_ALL_HD = tk.Button(self.root, text="Открыть файл со всеми считанными праздниками",
                                                  command=self.open_ALL_HD)
        self.open_button_ALL_HD.grid(row=2, column=0, columnspan = 1,  sticky="nsew")

        # Tables
        all_hd_label = tk.Label(self.root, text="История поиска:")
        all_hd_label.grid(row=3, column=0, columnspan=1, pady=10, sticky="nsew")

        self.all_hd_table = ttk.Treeview(self.root, columns=("Название", "Дата", "Тип"), show="headings")
        self.all_hd_table.heading("Название", text="Название События")
        self.all_hd_table.heading("Дата", text="Дата")
        self.all_hd_table.heading("Тип", text="Раздел")
        self.all_hd_table.grid(row=4, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")

        latest_hd_label = tk.Label(self.root, text="Итоги поиска:")
        latest_hd_label.grid(row=3, column=2, columnspan=2, pady=10, sticky="nsew")

        self.latest_hd_table = ttk.Treeview(self.root, columns=("Название", "Дата", "Тип", "Статус"),
                                                  show="headings")
        self.latest_hd_table.heading("Название", text="Название События")
        self.latest_hd_table.heading("Дата", text="Дата")
        self.latest_hd_table.heading("Тип", text="Раздел")
        self.latest_hd_table.heading("Статус", text="Статус")
        self.latest_hd_table.grid(row=4, column=3, columnspan=2, padx=10, pady=10, sticky="nsew")

    def parse_hd(self):
        section = self.section_var.get()
        day = self.day_var.get()
        month = self.month_var.get()




        base_url = sections[section]
        sheet_title = section
        today_hd_data = set()
        old_hd_data = set()
        all_hd_data = set()

        if os.path.exists(file_path):
            wb_old = load_workbook(file_path)
            if sheet_title in wb_old.sheetnames:
                ws_old_hd = wb_old[sheet_title]
                for row in ws_old_hd.iter_rows(min_row=2, values_only=True):
                    old_hd_data.add((row[0], row[1], row[2], row[3]))
            else:
                ws_old_hd = wb_old.create_sheet(sheet_title)
        else:
            wb_old = Workbook()
            ws_old_hd = wb_old.active
            ws_old_hd.title = sheet_title

        if ws_old_hd.max_row == 1:
            ws_old_hd.append(["Название события", "Ссылка на статью события", "Дата события", "Тип события", "Статус"])

        if os.path.exists(all_hd_file_path):
            wb_all_old = load_workbook(all_hd_file_path)
            if sheet_title in wb_all_old.sheetnames:
                ws_all_old_hd = wb_all_old[sheet_title]
                for row in ws_all_old_hd.iter_rows(min_row=2, values_only=True):
                    all_hd_data.add((row[0], row[1], row[2], row[3]))
            else:
                ws_all_old_hd = wb_all_old.create_sheet(sheet_title)
        else:
            wb_all_old = Workbook()
            ws_all_old_hd = wb_all_old.active
            ws_all_old_hd.title = sheet_title

        if ws_all_old_hd.max_row == 1:
            ws_all_old_hd.append(["Название события", "Ссылка на статью событи", "Дата событи", "Тип события"])

        current_date = datetime.now().strftime('%m-%d')
        montht1 = current_date[0:current_date.find('-')]
        mi = int(montht1)
        montht = str(mi)
        dayt1 = current_date[current_date.find('-') + 1:len(current_date)]
        di = int(dayt1)
        dayt = str(di)

        hd_links = collect_links(base_url, day, month)
        hd_data = collect_hd_data(hd_links,day,month)

        hd_linkst = collect_links(base_url, dayt, montht)
        hd_datat = collect_hd_data(hd_linkst, dayt, montht)

        save_data_to_sheet(ws_old_hd, hd_data, old_hd_data)
        wb_old.save(file_path)
        print(f"Праздники сохранены в файл {file_path}")

        for title, link, date, type_str in hd_data:
            if (title, link, date, type_str) not in all_hd_data:
                ws_all_old_hd.append([title, link, date, type_str])
                all_hd_data.add((title, link, date, type_str))
        wb_all_old.save(all_hd_file_path)
        print(f"Все события сохранены в файл {all_hd_file_path}")

        self.populate_all_hd_table(section)
        self.populate_latest_hd_table(section)


    def open_ALL_HD(self):
        os.startfile(all_hd_file_path)

    def show_hd(self):
        section = self.section_var.get()

        self.populate_all_hd_table(section)
        self.populate_latest_hd_table(section)

    def populate_all_hd_table(self, section):
        for i in self.all_hd_table.get_children():
            self.all_hd_table.delete(i)

        if os.path.exists(all_hd_file_path):
            wb = load_workbook(all_hd_file_path)
            if section in wb.sheetnames:
                            ws = wb[section]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                self.all_hd_table.insert("", "end", values=(row[0], row[2], row[3]))

    def populate_latest_hd_table(self, section):
                    for i in self.latest_hd_table.get_children():
                        self.latest_hd_table.delete(i)

                    if os.path.exists(file_path):
                        wb = load_workbook(file_path)
                        if section in wb.sheetnames:
                            ws = wb[section]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                self.latest_hd_table.insert("", "end", values=(row[0], row[2], row[3], row[4]))

if __name__ == "__main__":
    root = tk.Tk()
    app = HolidaysParserApp(root)
    root.mainloop()

